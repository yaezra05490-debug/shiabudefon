import streamlit as st
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
import google.generativeai as genai
import os
import re
import plotly.express as px
from io import BytesIO
from datetime import datetime
import functools

# ============================
# הגדרת הדף
# ============================
st.set_page_config(
    page_title="שיעבודא פון",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ============================
# CSS + JS גלובלי
# ============================
def inject_css():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Heebo:wght@300;400;500;600;700;800;900&display=swap');

    /* ── בסיס ── */
    html, body, [class*="css"], .stApp {
        direction: rtl;
        font-family: 'Heebo', 'Segoe UI', Arial, sans-serif !important;
    }
    .stApp { background: #f4f6fb !important; }
    #MainMenu, footer, .stDeployButton, [data-testid="stToolbar"] { visibility: hidden !important; display:none !important; }
    .block-container { padding-top: 0 !important; padding-bottom: 2rem !important; max-width: 1300px !important; }
    div[data-testid="stVerticalBlock"] > div { gap: 0 !important; }

    /* ── Header כפתורים ── */
    div[data-testid="stHorizontalBlock"] button {
        border-radius: 22px !important;
        font-family: 'Heebo', sans-serif !important;
        font-size: 0.85rem !important;
        font-weight: 500 !important;
        transition: all 0.2s ease !important;
        padding: 5px 16px !important;
    }
    div[data-testid="stHorizontalBlock"] button[kind="primary"] {
        background: #e8b84b !important;
        color: #1a1a2e !important;
        border: none !important;
        font-weight: 700 !important;
        box-shadow: 0 2px 8px rgba(232,184,75,0.4) !important;
    }
    div[data-testid="stHorizontalBlock"] button[kind="secondary"] {
        background: rgba(255,255,255,0.08) !important;
        color: #ccd6f6 !important;
        border: 1px solid rgba(255,255,255,0.2) !important;
    }
    div[data-testid="stHorizontalBlock"] button[kind="secondary"]:hover {
        background: rgba(232,184,75,0.15) !important;
        color: #e8b84b !important;
        border-color: #e8b84b !important;
    }

    /* ── Header ── */
    .header-wrap {
        background: linear-gradient(135deg, #0d1b2a 0%, #1a1a2e 40%, #0f3460 100%);
        padding: 0 24px;
        margin: 0 -1rem 0 -1rem;
        box-shadow: 0 4px 20px rgba(0,0,0,0.35);
        border-bottom: 2px solid rgba(232,184,75,0.3);
    }
    .header-logo {
        color: #e8b84b;
        font-size: 1.35rem;
        font-weight: 900;
        letter-spacing: 1.5px;
        padding: 14px 0;
        text-shadow: 0 2px 8px rgba(232,184,75,0.3);
    }
    .header-logo span { color: #fff; font-weight: 300; font-size: 1rem; margin-right: 4px; }

    /* ── Tabs ── */
    .stTabs [data-baseweb="tab-list"] {
        background: white !important;
        border-radius: 14px !important;
        padding: 6px !important;
        box-shadow: 0 2px 12px rgba(0,0,0,0.08) !important;
        gap: 4px !important;
        margin-bottom: 20px !important;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 10px !important;
        font-family: 'Heebo', sans-serif !important;
        font-size: 0.9rem !important;
        font-weight: 600 !important;
        color: #666 !important;
        padding: 8px 20px !important;
        border: none !important;
        background: transparent !important;
        transition: all 0.2s !important;
    }
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #0f3460, #1a237e) !important;
        color: white !important;
        box-shadow: 0 3px 10px rgba(15,52,96,0.3) !important;
    }
    .stTabs [data-baseweb="tab-highlight"] { display: none !important; }
    .stTabs [data-baseweb="tab-border"] { display: none !important; }

    /* ── Inputs ── */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stSelectbox > div > div > div {
        border-radius: 10px !important;
        border: 1.5px solid #e2e8f0 !important;
        font-family: 'Heebo', sans-serif !important;
        font-size: 0.92rem !important;
        padding: 8px 14px !important;
        direction: rtl !important;
        transition: border-color 0.2s !important;
    }
    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus {
        border-color: #0f3460 !important;
        box-shadow: 0 0 0 3px rgba(15,52,96,0.1) !important;
    }
    .stTextInput label, .stNumberInput label, .stSelectbox label, .stMultiselect label {
        font-family: 'Heebo', sans-serif !important;
        font-size: 0.82rem !important;
        font-weight: 600 !important;
        color: #475569 !important;
    }

    /* ── כפתורים כלליים ── */
    .stButton > button {
        border-radius: 10px !important;
        font-family: 'Heebo', sans-serif !important;
        font-weight: 600 !important;
        transition: all 0.2s !important;
    }
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #0f3460, #1565c0) !important;
        color: white !important;
        border: none !important;
        box-shadow: 0 3px 12px rgba(15,52,96,0.35) !important;
    }
    .stButton > button[kind="primary"]:hover {
        transform: translateY(-1px) !important;
        box-shadow: 0 5px 18px rgba(15,52,96,0.45) !important;
    }

    /* ── Login ── */
    .login-screen {
        position: fixed;
        top: 0; left: 0; right: 0; bottom: 0;
        background: linear-gradient(135deg, #0d1b2a 0%, #0f3460 50%, #1565c0 100%);
        display: flex;
        align-items: center;
        justify-content: center;
        z-index: 9999;
    }
    .login-card {
        background: white;
        border-radius: 24px;
        padding: 48px 44px 40px;
        box-shadow: 0 24px 80px rgba(0,0,0,0.35);
        width: 100%;
        max-width: 420px;
        direction: rtl;
        animation: fadeUp 0.4s ease;
    }
    @keyframes fadeUp {
        from { opacity:0; transform: translateY(20px); }
        to   { opacity:1; transform: translateY(0); }
    }
    .login-logo {
        text-align: center;
        font-size: 3rem;
        margin-bottom: 8px;
    }
    .login-title {
        text-align: center;
        font-size: 1.7rem;
        font-weight: 800;
        color: #0f3460;
        margin-bottom: 4px;
    }
    .login-sub {
        text-align: center;
        color: #94a3b8;
        font-size: 0.88rem;
        margin-bottom: 28px;
    }
    .login-card .stTextInput > div > div > input {
        background: #f8fafc !important;
    }
    .login-divider {
        text-align: center;
        color: #cbd5e1;
        font-size: 0.8rem;
        margin: 16px 0;
        position: relative;
    }
    .login-divider::before, .login-divider::after {
        content: '';
        position: absolute;
        top: 50%;
        width: 40%;
        height: 1px;
        background: #e2e8f0;
    }
    .login-divider::before { right: 0; }
    .login-divider::after  { left: 0; }

    /* ── Greeting bar ── */
    .greeting-bar {
        background: linear-gradient(135deg, #0f3460 0%, #1565c0 100%);
        border-radius: 16px;
        padding: 20px 28px;
        margin-bottom: 20px;
        display: flex;
        align-items: center;
        justify-content: space-between;
        color: white;
    }
    .greeting-name { font-size: 1.3rem; font-weight: 700; }
    .greeting-uid  { font-size: 0.82rem; opacity: 0.75; margin-top: 2px; }
    .greeting-last { font-size: 0.75rem; opacity: 0.8; margin-top: 6px; background: rgba(255,255,255,0.12); border-radius: 8px; padding: 3px 10px; display: inline-block; }
    .greeting-last b { font-weight: 700; opacity: 1; }
    .greeting-badge {
        background: rgba(255,255,255,0.15);
        border-radius: 12px;
        padding: 8px 18px;
        font-size: 0.82rem;
        font-weight: 600;
        border: 1px solid rgba(255,255,255,0.2);
    }

    /* ── Metric cards ── */
    .mc-wrap { border-radius: 18px; padding: 24px 26px; height: 100%; color: white; position: relative; overflow: hidden; }
    .mc-wrap::after {
        content: '';
        position: absolute;
        top: -20px; left: -20px;
        width: 100px; height: 100px;
        border-radius: 50%;
        background: rgba(255,255,255,0.08);
    }
    .mc-blue   { background: linear-gradient(135deg, #0f3460, #1565c0); box-shadow: 0 6px 20px rgba(15,52,96,0.35); }
    .mc-green  { background: linear-gradient(135deg, #1b5e20, #388e3c); box-shadow: 0 6px 20px rgba(27,94,32,0.35); }
    .mc-purple { background: linear-gradient(135deg, #4a148c, #7b1fa2); box-shadow: 0 6px 20px rgba(74,20,140,0.35); }
    .mc-label  { font-size: 0.78rem; opacity: 0.8; margin-bottom: 8px; font-weight: 500; }
    .mc-value  { font-size: 2.1rem; font-weight: 800; line-height: 1; margin-bottom: 8px; }
    .mc-sub    { font-size: 0.7rem; opacity: 0.65; }
    .mc-icon   { position: absolute; top: 18px; left: 20px; font-size: 2rem; opacity: 0.2; }

    /* ── Smart cards ── */
    .smart-card {
        border-radius: 14px;
        padding: 14px 18px;
        margin-bottom: 10px;
        display: flex;
        align-items: center;
        gap: 14px;
        box-shadow: 0 2px 12px rgba(0,0,0,0.06);
        direction: rtl;
        transition: transform 0.15s, box-shadow 0.15s;
        border: 1px solid transparent;
    }
    .smart-card:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(0,0,0,0.1); }
    .smart-card.debit  { background: #fff5f5; border-right: 4px solid #e53935; border-color: #fee2e2; }
    .smart-card.credit { background: #f0fdf4; border-right: 4px solid #43a047; border-color: #dcfce7; }
    .smart-card.failed { background: #fffbeb; border-right: 4px solid #f9a825; border-color: #fef3c7; }
    .smart-card.admin  { background: #eff6ff; border-right: 4px solid #1e88e5; border-color: #dbeafe; }
    .card-circle {
        min-width: 60px; height: 60px;
        border-radius: 14px;
        display: flex; align-items: center; justify-content: center;
        font-weight: 800; font-size: 0.72rem;
        flex-shrink: 0; text-align: center; line-height: 1.3;
        direction: ltr;
    }
    .debit  .card-circle { background: #fee2e2; color: #b71c1c; }
    .credit .card-circle { background: #dcfce7; color: #14532d; }
    .failed .card-circle { background: #fef3c7; color: #92400e; }
    .admin  .card-circle { background: #dbeafe; color: #1e3a8a; }
    .card-main { flex: 1; min-width: 0; direction: rtl; }
    .card-main .c-title {
        font-weight: 600; font-size: 0.9rem; color: #1e293b; margin-bottom: 4px;
        white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
        direction: rtl; text-align: right;
    }
    .card-main .c-sub   { font-size: 0.78rem; color: #64748b; direction: rtl; text-align: right; }
    .card-direction {
        font-size: 0.7rem; font-weight: 700; padding: 2px 8px;
        border-radius: 8px; margin-top: 4px; display: inline-block;
    }
    .dir-credit { background: #dcfce7; color: #14532d; }
    .dir-debit  { background: #fee2e2; color: #b71c1c; }
    .dir-admin  { background: #dbeafe; color: #1e3a8a; }
    .card-side { text-align: right; min-width: 100px; flex-shrink: 0; direction: rtl; }
    .badge { display: inline-block; padding: 3px 10px; border-radius: 20px; font-size: 0.68rem; font-weight: 700; margin-bottom: 4px; }
    .badge-ok     { background: #dcfce7; color: #14532d; }
    .badge-failed { background: #fef3c7; color: #92400e; }
    .badge-admin  { background: #dbeafe; color: #1e3a8a; }
    .c-date  { font-size: 0.7rem; color: #94a3b8; margin: 3px 0; direction: ltr; text-align: right; }
    .c-phone { font-size: 0.68rem; color: #cbd5e1; direction: ltr; text-align: right; }

    /* ── Section titles ── */
    .section-title {
        font-size: 1rem;
        font-weight: 700;
        color: #0f3460;
        margin: 22px 0 12px;
        padding: 10px 14px;
        background: white;
        border-radius: 10px;
        border-right: 4px solid #e8b84b;
        box-shadow: 0 1px 6px rgba(0,0,0,0.05);
    }

    /* ── Info pages ── */
    .info-page {
        max-width: 820px; margin: 30px auto;
        background: white; border-radius: 20px;
        padding: 44px 52px;
        box-shadow: 0 4px 24px rgba(0,0,0,0.08);
        line-height: 2;
    }
    .info-page h2 { color: #0f3460; border-bottom: 2px solid #e8b84b; padding-bottom: 10px; margin-bottom: 20px; }

    /* ── Personal details ── */
    .detail-block {
        background: white; border-radius: 16px;
        padding: 28px 32px;
        box-shadow: 0 2px 14px rgba(0,0,0,0.06);
        margin-bottom: 18px;
    }
    .detail-row { display: flex; gap: 16px; margin-bottom: 14px; align-items: center; flex-wrap: wrap; border-bottom: 1px solid #f1f5f9; padding-bottom: 10px; }
    .detail-row:last-of-type { border-bottom: none; margin-bottom: 0; padding-bottom: 0; }
    .detail-label { color: #64748b; font-size: 0.82rem; min-width: 130px; font-weight: 500; }
    .detail-value { font-weight: 700; color: #0f172a; font-size: 0.95rem; }
    .phone-tag {
        display: inline-block;
        background: linear-gradient(135deg, #eff6ff, #dbeafe);
        color: #1d4ed8;
        border-radius: 20px;
        padding: 5px 16px;
        font-size: 0.85rem;
        margin: 4px;
        font-weight: 600;
        border: 1px solid #bfdbfe;
    }
    .readonly-note {
        background: #fffbeb;
        border: 1px solid #fde68a;
        border-radius: 12px;
        padding: 12px 16px;
        font-size: 0.82rem;
        color: #92400e;
        margin-top: 16px;
        display: flex;
        gap: 8px;
        align-items: flex-start;
    }

    /* ── Floating chat ── */
    #chat-fab-wrapper { position: fixed !important; bottom: 28px !important; left: 28px !important; z-index: 99999 !important; }
    .chat-fab-btn {
        width: 60px; height: 60px;
        border-radius: 50%;
        background: linear-gradient(135deg, #0f3460, #1565c0);
        border: none; cursor: pointer;
        font-size: 1.5rem; color: white;
        box-shadow: 0 6px 24px rgba(15,52,96,0.45);
        animation: pulse 2.5s infinite;
        display: flex; align-items: center; justify-content: center;
        transition: transform 0.2s;
    }
    .chat-fab-btn:hover { transform: scale(1.08); }
    .chat-panel-fixed {
        position: fixed !important;
        bottom: 98px !important;
        left: 28px !important;
        width: 360px;
        max-height: 480px;
        background: white;
        border-radius: 20px;
        box-shadow: 0 12px 40px rgba(0,0,0,0.2);
        z-index: 99998 !important;
        flex-direction: column;
        overflow: hidden;
        border: 1px solid #e2e8f0;
    }
    .chat-panel-hdr {
        background: linear-gradient(135deg, #0f3460, #1565c0);
        color: white;
        padding: 16px 20px;
        font-weight: 700;
        font-size: 0.95rem;
        display: flex;
        justify-content: space-between;
        align-items: center;
    }
    @keyframes pulse {
        0%   { box-shadow: 0 0 0 0 rgba(15,52,96,0.5); }
        70%  { box-shadow: 0 0 0 14px rgba(15,52,96,0); }
        100% { box-shadow: 0 0 0 0 rgba(15,52,96,0); }
    }

    /* ── Divider ── */
    hr { border: none !important; border-top: 1.5px solid #e2e8f0 !important; margin: 20px 0 !important; }

    /* ── Radio buttons ── */
    .stRadio > div {
        background: white;
        border-radius: 12px;
        padding: 4px 8px;
        box-shadow: 0 1px 6px rgba(0,0,0,0.05);
        display: inline-flex;
        gap: 4px;
    }
    .stRadio [data-testid="stMarkdownContainer"] p { font-family: 'Heebo', sans-serif !important; font-size: 0.84rem !important; font-weight: 600 !important; }

    /* ── Multiselect tags ── */
    [data-baseweb="tag"] {
        background: linear-gradient(135deg,#dbeafe,#bfdbfe) !important;
        border-radius: 20px !important;
        border: none !important;
    }
    [data-baseweb="tag"] span { color: #1e3a8a !important; font-weight: 600 !important; font-size: 0.8rem !important; }

    /* ── Download buttons ── */
    [data-testid="stDownloadButton"] > button {
        border-radius: 10px !important;
        font-family: 'Heebo', sans-serif !important;
        font-weight: 600 !important;
        font-size: 0.84rem !important;
        background: white !important;
        border: 1.5px solid #e2e8f0 !important;
        color: #0f3460 !important;
        transition: all 0.18s !important;
    }
    [data-testid="stDownloadButton"] > button:hover {
        background: #eff6ff !important;
        border-color: #93c5fd !important;
        transform: translateY(-1px) !important;
        box-shadow: 0 3px 10px rgba(15,52,96,0.15) !important;
    }

    /* ── Chat RTL ── */
    [data-testid="stChatMessage"] { direction: rtl !important; text-align: right !important; }
    [data-testid="stChatMessage"] > div { direction: rtl !important; }
    [data-testid="stChatMessage"] p { text-align: right !important; }
    .stChatInputContainer textarea {
        direction: rtl !important; text-align: right !important;
        font-family: 'Heebo', sans-serif !important;
    }

    /* ── Pagination ── */
    .pg-info {
        background: white; border-radius: 12px; padding: 10px 20px;
        font-size: 0.84rem; font-weight: 600; color: #475569;
        box-shadow: 0 1px 8px rgba(0,0,0,0.06); text-align: center;
    }

    /* ── Spinner ── */
    [data-testid="stSpinner"] p { font-family: 'Heebo', sans-serif !important; color: #475569 !important; }

    /* ── Caption / small text ── */
    .stCaption, small { color: #94a3b8 !important; font-size: 0.78rem !important; }

    /* ── Mobile responsive ── */
    @media (max-width: 768px) {
        .smart-card { flex-direction: column; align-items: flex-start; }
        .card-side  { text-align: right; min-width: unset; }
        .detail-row { flex-direction: column; gap: 4px; }
        .chat-panel-fixed { width: calc(100vw - 60px) !important; left: 10px !important; }
        .block-container { padding-left: 8px !important; padding-right: 8px !important; }
        .mc-value { font-size: 1.5rem !important; }
        .greeting-bar { flex-direction: column; gap: 10px; text-align: center; }
        /* hamburger: hide desktop nav column, show hamburger */
        .hamburger-btn { display: flex !important; }
        /* hide the 2nd column in header (nav buttons) */
        [data-testid="stHorizontalBlock"]:has(.nav-desktop-marker) > [data-testid="column"]:nth-child(2) > div {
            display: none !important;
        }
    }
    @media (min-width: 769px) {
        .hamburger-btn { display: none !important; }
        .mobile-nav-drawer { display: none !important; }
    }
    .hamburger-btn {
        background: none; border: none; cursor: pointer;
        padding: 10px 8px; flex-direction: column; gap: 5px; align-items: center; justify-content: center;
    }
    .hamburger-btn span { display: block; width: 24px; height: 2px; background: #e8b84b; border-radius: 2px; transition: all 0.2s; }
    .mobile-nav-drawer {
        position: fixed; top: 0; right: 0; width: 240px; height: 100vh;
        background: linear-gradient(180deg,#0d1b2a,#0f3460);
        z-index: 99999; padding: 60px 0 20px;
        box-shadow: -4px 0 24px rgba(0,0,0,0.4);
        overflow-y: auto; direction: rtl;
        transform: translateX(240px);
        transition: transform 0.3s ease;
    }
    .mobile-nav-drawer a {
        display: block; padding: 14px 24px; color: #ccd6f6;
        text-decoration: none; font-family: 'Heebo', sans-serif;
        font-size: 1rem; font-weight: 600; border-bottom: 1px solid rgba(255,255,255,0.07);
        transition: background 0.15s;
    }
    .mobile-nav-drawer a:hover, .mobile-nav-drawer a.active { background: rgba(232,184,75,0.15); color: #e8b84b; }
    .mobile-nav-overlay {
        display: none; position: fixed; inset: 0;
        background: rgba(0,0,0,0.5); z-index: 99998;
    }
    </style>
    """, unsafe_allow_html=True)


def inject_chat_fab_js():
    """FIX #6 – כפתור צף אמיתי עם JS."""
    st.markdown("""
    <div id="chat-fab-wrapper">
        <button class="chat-fab-btn" onclick="toggleChatPanel()" id="chat-fab" title="בוט שירות">💬</button>
    </div>
    <div class="chat-panel-fixed" id="chat-panel" style="display:none;">
        <div class="chat-panel-hdr">
            <span>🤖 בוט שיעבודא פון</span>
            <button onclick="toggleChatPanel()" style="background:none;border:none;color:white;cursor:pointer;font-size:1.1rem;">✕</button>
        </div>
        <div style="padding:14px;flex:1;overflow-y:auto;font-size:0.85rem;color:#555;text-align:right;direction:rtl">
            <p>שלום! אני הבוט של שיעבודא פון 👋<br>לשיחה מלאה עם ה-AI, כנס לכרטיסיית <b>🤖 צ'אט AI</b>.</p>
        </div>
    </div>
    <script>
    function toggleChatPanel() {
        var panel = document.getElementById('chat-panel');
        var fab   = document.getElementById('chat-fab');
        if (panel.style.display === 'none' || panel.style.display === '') {
            panel.style.display = 'flex';
            fab.innerHTML = '✕';
        } else {
            panel.style.display = 'none';
            fab.innerHTML = '💬';
        }
    }
    </script>
    """, unsafe_allow_html=True)


inject_css()


def style_fig(fig, height=300):
    """תבנית Plotly אחידה."""
    fig.update_layout(
        height=height,
        margin=dict(l=0, r=0, t=30, b=0),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="#fafafa",
        font=dict(family="Heebo, Arial", size=12, color="#475569"),
        xaxis=dict(tickangle=-45, gridcolor="#f1f5f9", showgrid=True,
                   linecolor="#e2e8f0", tickfont=dict(size=11)),
        yaxis=dict(gridcolor="#f1f5f9", showgrid=True,
                   linecolor="#e2e8f0", tickfont=dict(size=11)),
        legend=dict(orientation="h", yanchor="bottom", y=1.02,
                    xanchor="right", x=1, font=dict(size=11)),
        hoverlabel=dict(bgcolor="white", font_size=13,
                        font_family="Heebo, Arial", bordercolor="#e2e8f0"),
    )
    fig.update_traces(marker_line_width=0)
    return fig


def render_empty_state(icon: str, title: str, subtitle: str = ""):
    """כרטיס empty state מעוצב."""
    sub_html = f'<div style="font-size:0.84rem;color:#94a3b8;margin-top:4px">{subtitle}</div>' if subtitle else ""
    st.markdown(f"""
    <div style="text-align:center;padding:44px 24px;background:white;border-radius:18px;
                box-shadow:0 2px 12px rgba(0,0,0,0.06);margin:8px 0">
        <div style="font-size:3.2rem;margin-bottom:14px">{icon}</div>
        <div style="font-size:1rem;font-weight:700;color:#1e293b">{title}</div>
        {sub_html}
    </div>""", unsafe_allow_html=True)


# ============================
# משתנים
# ============================
SPREADSHEET_ID = '1PB-FJsvBmCy8hwA_S1S5FLY_QU9P2VstDAJMMdtufHM'
SYSTEM_MANUAL = """
אתה הנציג הדיגיטלי של שיעבודא פון. תפקידך לשמש כמנתח נתונים של המשתמש.
מותר לך לסכם לו את הפעולות, לחשב הוצאות/הכנסות, ולענות על שאלות בצורה אדיבה ומקצועית.
אל תחשוף מידע על משתמשים אחרים. ענה תמיד בעברית.
"""

# ============================
# Backend
# ============================
@st.cache_resource
def get_client():
    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    secret_dir = "/etc/secrets"
    required_keys = ["type","project_id","private_key_id","private_key","client_email",
                     "client_id","auth_uri","token_uri","auth_provider_x509_cert_url","client_x509_cert_url"]
    errors = []

    # ── שיטה 1: st.secrets["gcp_service_account"] (Streamlit Cloud / Render secrets.toml) ──
    try:
        raw = st.secrets["gcp_service_account"]
        creds_dict = dict(raw)
        if "private_key" in creds_dict:
            creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n")
        return gspread.authorize(Credentials.from_service_account_info(creds_dict, scopes=scopes))
    except (KeyError, FileNotFoundError):
        errors.append("st.secrets['gcp_service_account'] – לא נמצא")
    except Exception as e:
        errors.append(f"st.secrets['gcp_service_account'] – שגיאה: {e}")

    # ── שיטה 2: st.secrets["google_credentials"] (מפתח חלופי) ──
    try:
        import json as _json
        raw_str = st.secrets["google_credentials"]
        creds_dict = _json.loads(raw_str) if isinstance(raw_str, str) else dict(raw_str)
        if "private_key" in creds_dict:
            creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n")
        return gspread.authorize(Credentials.from_service_account_info(creds_dict, scopes=scopes))
    except (KeyError, FileNotFoundError):
        errors.append("st.secrets['google_credentials'] – לא נמצא")
    except Exception as e:
        errors.append(f"st.secrets['google_credentials'] – שגיאה: {e}")

    # ── שיטה 3: קבצים בודדים תחת /etc/secrets/{key} (Render Secret Files) ──
    try:
        creds_dict = {}
        for key in required_keys:
            path = os.path.join(secret_dir, key)
            if os.path.exists(path):
                with open(path, "r") as f:
                    val = f.read().strip()
                    if key == "private_key":
                        val = val.replace("\\n", "\n").replace('"', '')
                    creds_dict[key] = val
        if "private_key" in creds_dict and "client_email" in creds_dict:
            return gspread.authorize(Credentials.from_service_account_info(creds_dict, scopes=scopes))
        else:
            errors.append(f"קבצים בודדים /etc/secrets/ – חסרים מפתחות ({', '.join(k for k in required_keys if k not in creds_dict)})")
    except Exception as e:
        errors.append(f"קבצים בודדים /etc/secrets/ – שגיאה: {e}")

    # ── שיטה 4: קובץ service_account.json ──
    json_path = os.path.join(secret_dir, "service_account.json")
    if os.path.exists(json_path):
        try:
            return gspread.authorize(Credentials.from_service_account_file(json_path, scopes=scopes))
        except Exception as e:
            errors.append(f"service_account.json – שגיאה: {e}")
    else:
        errors.append(f"service_account.json – לא נמצא ב-{json_path}")

    # ── כישלון מלא ──
    st.error("❌ **תקלה בחיבור ל-Google Sheets** – לא נמצאו קרדנשיאלים תקינים.\n\n" +
             "\n\n".join(f"• {e}" for e in errors) +
             "\n\n**כיצד לתקן:** הוסף ב-Render → Environment → Secret Files את הקובץ `service_account.json`, "
             "או הגדר ב-Render → Environment Variables → `gcp_service_account` עם תוכן ה-JSON.")
    st.stop()


@st.cache_resource
def get_api_keys():
    keys = []
    try:
        if "gemini_api_key" in st.secrets:
            val = st.secrets["gemini_api_key"]
            raw = val.get("api_keys", val.get("api_key","")) if isinstance(val,dict) else str(val)
            keys = [k.strip() for k in raw.split(",") if k.strip()]
    except Exception:
        pass
    if not keys:
        for path in ["/etc/secrets/api_key","/etc/secrets/gemini_api_key"]:
            if os.path.exists(path):
                with open(path,"r") as f:
                    keys = [k.strip() for k in f.read().strip().replace('"','').replace("'","").split(",") if k.strip()]
                break
    return keys


@st.cache_data(ttl=3600)
def _get_gemini_model_name():
    """Cache the best Gemini model name — queried once per hour."""
    try:
        best = "gemini-pro"
        for m in genai.list_models():
            if "generateContent" in m.supported_generation_methods:
                best = m.name
                if "flash" in m.name.lower():
                    break
        return best
    except Exception:
        return "gemini-pro"


@st.cache_data(ttl=600)
def get_all_data():
    """FIX #13 – cache 10 דקות."""
    client = get_client()
    sh = client.open_by_key(SPREADSHEET_ID)
    df_users   = pd.DataFrame(sh.worksheet("משתמשים").get_all_records())
    df_actions = pd.DataFrame(sh.worksheet("פעולות").get_all_records())
    try:
        admin_ids = pd.DataFrame(sh.worksheet("מנהלים").get_all_records()).iloc[:,0].astype(str).tolist()
    except Exception:
        admin_ids = []
    return df_users, df_actions, admin_ids


@st.cache_data(ttl=600)
def get_cms_content():
    defaults = {
        "קצת עלינו": "**שיעבודא פון** – מערכת ניהול חשבונות מתקדמת.\n\nאנחנו מספקים שירות מהיר, בטוח ושקוף.",
        "צור קשר":   "לפניות ותמיכה:\n📞 054-0000000\n📧 support@shiabudefon.com",
        "תקנון":     "השימוש במערכת כפוף לתנאי השימוש. כל הפעולות מבוצעות בצורה מאובטחת.",
    }
    try:
        sh  = get_client().open_by_key(SPREADSHEET_ID)
        df  = pd.DataFrame(sh.worksheet("תוכן").get_all_records())
        for _, row in df.iterrows():
            k,v = str(row.get("כותרת","")).strip(), str(row.get("תוכן","")).strip()
            if k and v: defaults[k] = v
    except Exception:
        pass
    return defaults


# ============================
# עיבוד נתונים
# ============================
def process_user_actions(df_actions: pd.DataFrame, user_id: str) -> pd.DataFrame:
    if df_actions.empty: return pd.DataFrame()
    df = df_actions.copy()
    df["מספר משתמש מקור"] = df["מספר משתמש מקור"].astype(str)
    df["מספר משתמש יעד"]  = df["מספר משתמש יעד"].astype(str)
    uid  = str(user_id)
    mask = (df["מספר משתמש מקור"] == uid) | (df["מספר משתמש יעד"] == uid)
    my   = df[mask].copy()
    if my.empty: return pd.DataFrame()

    def _s(v):
        s = str(v).strip()
        return s if s not in ('', 'nan', 'None', '0') else ''

    def enrich(row):
        is_sender = str(row["מספר משתמש מקור"]) == uid
        try:    amount = float(row.get("סכום",0))
        except: amount = 0
        direction = "חובה" if is_sender else "זכות"
        net  = -amount if is_sender else amount
        if is_sender:
            name, uid_t = _s(row.get('שם יעד','')), _s(row.get('מספר משתמש יעד',''))
            if name and uid_t:   desc = f"העברה אל {name} ({uid_t})"
            elif name:           desc = f"העברה אל {name}"
            elif uid_t:          desc = f"העברה אל משתמש {uid_t}"
            else:                desc = "העברה יוצאת"
        else:
            name, uid_s = _s(row.get('שם מקור','')), _s(row.get('מספר משתמש מקור',''))
            if name and uid_s:   desc = f"התקבל מ-{name} ({uid_s})"
            elif name:           desc = f"התקבל מ-{name}"
            elif uid_s:          desc = f"התקבל ממשתמש {uid_s}"
            else:                desc = "קבלת העברה"
        return pd.Series({"כיוון": direction, "סכום נטו": net, "תיאור": desc})

    enriched = my.apply(enrich, axis=1)
    return pd.concat([my.reset_index(drop=True), enriched.reset_index(drop=True)], axis=1)


def get_user_balance(df_users: pd.DataFrame, user_id: str) -> float:
    df = df_users.copy()
    df["מספר משתמש"] = df["מספר משתמש"].astype(str).str.strip()
    row = df[df["מספר משתמש"] == str(user_id)]
    if row.empty: return 0.0
    try:    return float(row.iloc[0]["יתרה"])
    except: return 0.0


# ============================
# FIX #8: כרטיסייה חכמה
# ============================
def render_smart_card(row: pd.Series, highlight: str = ""):
    status    = str(row.get("סטטוס","מוצלחת")).strip()
    direction = str(row.get("כיוון","")).strip()

    is_admin_op = "מנהל" in status
    is_failed   = "כושל" in status or "נכשל" in status

    if is_admin_op:
        cls, badge_cls, badge_txt = "admin",  "badge-admin",  "פעולת מנהל"
        dir_cls, dir_lbl = "dir-admin", "🛡️ פעולת מנהל"
    elif is_failed:
        cls, badge_cls, badge_txt = "failed", "badge-failed", "כושלת"
        dir_cls, dir_lbl = "dir-debit",  "❌ כושלת"
    elif direction == "זכות":
        cls, badge_cls, badge_txt = "credit", "badge-ok",     "מוצלחת"
        dir_cls, dir_lbl = "dir-credit", "⬆️ זכות"
    else:
        cls, badge_cls, badge_txt = "debit",  "badge-ok",     "מוצלחת"
        dir_cls, dir_lbl = "dir-debit",  "⬇️ חובה"

    try:    amount_str = f"₪{abs(float(row.get('סכום', 0))):,.0f}"
    except: amount_str = "₪-"

    # תיאור נקי: הסר מספרי סוגריים לשורה נפרדת
    raw_desc = str(row.get("תיאור", row.get("הערה", "")))
    uid_part = re.search(r'\(([^)]+)\)', raw_desc)
    uid_badge = f'<span style="direction:ltr;background:#f1f5f9;border-radius:6px;padding:1px 7px;font-size:0.7rem;color:#64748b;font-family:monospace;margin-right:6px">{uid_part.group(1)}</span>' if uid_part else ""
    clean_desc = re.sub(r'\s*\([^)]+\)', '', raw_desc).strip()
    if highlight.strip():
        hl = re.escape(highlight.strip())
        clean_desc = re.sub(f'({hl})', r'<mark style="background:#fef9c3;padding:0 2px;border-radius:3px">\1</mark>',
                            clean_desc, flags=re.IGNORECASE)

    after_raw = row.get("יתרה לאחר פעולה", row.get("יתרה", ""))
    date_val  = str(row.get("תאריך", ""))
    time_val  = str(row.get("שעה",   ""))
    phone     = str(row.get("טלפון", row.get("מבצע", "")))

    try:    after_str = f"₪{float(after_raw):,.2f}" if str(after_raw).strip() not in ("","nan","None") else ""
    except: after_str = ""

    after_html = f'<div class="c-sub">יתרה לאחר: <span style="direction:ltr;display:inline-block">{after_str}</span></div>' if after_str else ""
    phone_html = f'<div class="c-phone"><span style="direction:ltr;display:inline-block">{phone}</span></div>' if phone.strip() and phone != "nan" else ""
    time_html  = f'{date_val}<br><span style="color:#cbd5e1">{time_val}</span>' if time_val.strip() else date_val

    st.markdown(f"""
    <div class="smart-card {cls}">
        <div class="card-circle">{amount_str}</div>
        <div class="card-main">
            <div class="c-title">{clean_desc}</div>
            <div>{uid_badge}<span class="card-direction {dir_cls}">{dir_lbl}</span></div>
            {after_html}
        </div>
        <div class="card-side">
            <span class="badge {badge_cls}">{badge_txt}</span>
            <div class="c-date">{time_html}</div>
            {phone_html}
        </div>
    </div>""", unsafe_allow_html=True)


# ============================
# Header  FIX #12
# ============================
def render_header(current_page: str):
    pages = [
        ("app",     "🏠 אזור אישי" if st.session_state.get("authenticated") else "🏠 ראשי"),
        ("about",   "קצת עלינו"),
        ("contact", "צור קשר"),
        ("terms",   "תקנון"),
    ]
    st.markdown("""
    <style>
    .header-wrap { background:linear-gradient(135deg,#0d1b2a 0%,#1a1a2e 40%,#0f3460 100%);
        padding:0 28px; margin:-1rem -1rem 0 -1rem;
        box-shadow:0 4px 20px rgba(0,0,0,0.35);
        border-bottom:2px solid rgba(232,184,75,0.25); }
    </style>
    <div class="header-wrap"></div>""", unsafe_allow_html=True)

    with st.container():
        st.markdown(
            '<div style="background:linear-gradient(135deg,#0d1b2a,#0f3460);padding:0 28px;margin:-0.5rem -1rem 0;">',
            unsafe_allow_html=True)
        c_logo, c_nav, c_auth, c_hamburger = st.columns([2, 6, 2, 1])
        with c_logo:
            st.markdown('<div class="header-logo">💰 שיעבודא פון<span>|</span></div>',
                        unsafe_allow_html=True)
        with c_nav:
            st.markdown('<span class="nav-desktop-marker"></span>', unsafe_allow_html=True)
            nav_cols = st.columns(len(pages))
            for i, (page_key, label) in enumerate(pages):
                with nav_cols[i]:
                    t = "primary" if current_page == page_key else "secondary"
                    if st.button(label, key=f"nav_{page_key}", type=t, use_container_width=True):
                        st.session_state.page = page_key; st.rerun()
        with c_auth:
            if st.session_state.get("authenticated"):
                if st.button("🚪 יציאה", type="secondary", use_container_width=True):
                    st.session_state.authenticated = False
                    st.session_state.page = "login"; st.rerun()
            else:
                if st.button("🔑 כניסה", type="primary", use_container_width=True):
                    st.session_state.page = "login"; st.rerun()
        with c_hamburger:
            st.markdown("""
            <button class="hamburger-btn" onclick="openDrawer()" aria-label="תפריט">
                <span></span><span></span><span></span>
            </button>
            <div class="mobile-nav-overlay" id="nav-overlay" onclick="closeDrawer()"></div>
            <div class="mobile-nav-drawer" id="nav-drawer">
                <button onclick="closeDrawer()" style="position:absolute;top:16px;left:16px;background:none;border:none;color:#e8b84b;font-size:1.4rem;cursor:pointer">✕</button>
                <a href="?page=app">🏠 אזור אישי</a>
                <a href="?page=about">קצת עלינו</a>
                <a href="?page=contact">צור קשר</a>
                <a href="?page=terms">תקנון</a>
            </div>
            <script>
            function openDrawer(){
                document.getElementById('nav-drawer').style.transform='translateX(0)';
                document.getElementById('nav-overlay').style.display='block';
            }
            function closeDrawer(){
                document.getElementById('nav-drawer').style.transform='translateX(240px)';
                document.getElementById('nav-overlay').style.display='none';
            }
            // Handle ?page= query param from drawer links
            (function(){
                var p=new URLSearchParams(window.location.search).get('page');
                if(p) window.history.replaceState({},'',window.location.pathname);
            })();
            </script>
            """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)


# ============================
# FIX #1 + #2: Login
# ============================
def render_login_page():
    st.markdown("""
    <style>
    /* Login: blue background full page */
    .block-container {
        background: linear-gradient(135deg, #0d1b2a 0%, #0f3460 55%, #1565c0 100%) !important;
        min-height: 88vh !important;
        padding: 60px 20px 40px !important;
        border-radius: 0 !important;
        max-width: 100% !important;
    }
    /* center column as white card */
    [data-testid="stColumns"] > [data-testid="column"]:nth-child(2) > div:first-child > div:first-child {
        background: white !important;
        border-radius: 24px !important;
        padding: 36px 40px 32px !important;
        box-shadow: 0 24px 80px rgba(0,0,0,0.45) !important;
        animation: fadeUp 0.35s ease;
    }
    </style>
    """, unsafe_allow_html=True)

    _, col_form, _ = st.columns([1, 1.3, 1])
    with col_form:
        st.markdown("""
        <div style="text-align:center;padding-bottom:4px">
            <div style="font-size:3rem;line-height:1.1">💰</div>
            <div style="font-size:1.75rem;font-weight:900;color:#0f3460;margin:10px 0 4px">שיעבודא פון</div>
            <div style="color:#94a3b8;font-size:0.88rem;margin-bottom:28px">ניהול חשבון אישי &middot; כניסה מאובטחת</div>
        </div>""", unsafe_allow_html=True)

        # FIX #1 – כפתור עין
        if "show_pwd" not in st.session_state:
            st.session_state.show_pwd = False

        uid = st.text_input("👤 מספר משתמש", placeholder="הזן מספר משתמש", key="login_uid")

        pwd_col, eye_col = st.columns([10, 1])
        with pwd_col:
            pwd = (st.text_input("🔓 סיסמה (מוצגת)",  value=st.session_state.get("_pwd_val",""), key="pwd_vis")
                   if st.session_state.show_pwd
                   else st.text_input("🔒 סיסמה", type="password", key="pwd_hid"))
            st.session_state._pwd_val = pwd
        with eye_col:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🙈" if st.session_state.show_pwd else "👁", key="eye", help="הצג/הסתר"):
                st.session_state.show_pwd = not st.session_state.show_pwd; st.rerun()

        if st.button("🔓 התחבר", use_container_width=True, type="primary", key="login_btn"):
            _do_login(uid, pwd)

        st.caption("אין הרשמה עצמאית. לפרטים חייג למערכת הטלפונית.")


def _do_login(uid: str, pwd: str):
    """Login with rate limiting, digit validation, unified error message."""
    TIMEOUT_SECS = 15 * 60  # 15 min
    MAX_ATTEMPTS = 5
    LOCKOUT_SECS = 5 * 60   # 5 min lockout

    # Rate limiting
    now = datetime.now().timestamp()
    attempts = st.session_state.get("_login_attempts", 0)
    lockout_until = st.session_state.get("_lockout_until", 0)
    if now < lockout_until:
        remaining = int((lockout_until - now) / 60) + 1
        st.error(f"🔒 חשבון חסום זמנית עקב ניסיונות כושלים מבו. נסה שוב בעוד {remaining} דקות.")
        return

    if not uid.strip(): st.error("⚠️ נא להזין מספר משתמש"); return
    if not uid.strip().isdigit(): st.error("⚠️ מספר משתמש חייב להכיל ספרות בלבד"); return
    if not pwd.strip(): st.error("⚠️ נא להזין סיסמה"); return
    try:
        with st.spinner("מתחבר..."):
            df_users, _, admin_ids = get_all_data()
        df_users["מספר משתמש"] = df_users["מספר משתמש"].astype(str).str.strip()
        df_users["סיסמה"]       = df_users["סיסמה"].astype(str).str.strip()
        uid_c, pwd_c = uid.strip(), pwd.strip()
        user_row = df_users[df_users["מספר משתמש"] == uid_c]
        # Unified error — don't reveal which field is wrong
        if user_row.empty or user_row.iloc[0]["סיסמה"] != pwd_c:
            attempts += 1
            st.session_state._login_attempts = attempts
            if attempts >= MAX_ATTEMPTS:
                st.session_state._lockout_until = now + LOCKOUT_SECS
                st.session_state._login_attempts = 0
                st.error("🔒 יותר מדי ניסיונות כושלים. החשבון נחסם ל-5 דקות.")
            else:
                left = MAX_ATTEMPTS - attempts
                st.error(f"❌ פרטי ההתחברות שגויים, אנא נסה שוב. ניסיונות נותרות: {left}")
            return
        # Success
        st.session_state._login_attempts = 0
        st.session_state._lockout_until  = 0
        st.session_state.authenticated = True
        st.session_state.user     = user_row.iloc[0].to_dict()
        st.session_state.is_admin = uid_c in [str(x).strip() for x in admin_ids]
        st.session_state.page     = "app"
        st.session_state._last_activity = datetime.now().timestamp()
        st.rerun()
    except Exception as e:
        st.error("המערכת מתעדכנת כעת, אנא המתן כדקה ונסה שוב.")


# ============================
# Tab 1 – דשבורד
# ============================
def render_dashboard(u, is_admin, df_users, df_actions):
    uid        = str(u.get("מספר משתמש",""))
    balance    = get_user_balance(df_users, uid)
    sync_time  = st.session_state.get("_last_sync", datetime.now().strftime("%H:%M:%S"))
    my_act     = process_user_actions(df_actions, uid)

    # פעולה אחרונה לתצוגה בברכה
    last_act_html = ""
    if not my_act.empty:
        try:
            lr = my_act.iloc[-1]
            la_icon = "📥" if str(lr.get("כיוון","")) == "זכות" else "📤"
            la_amt  = f"₪{abs(float(lr.get('סכום',0))):,.0f}"
            la_date = str(lr.get("תאריך",""))
            last_act_html = f'<div class="greeting-last">{la_icon} פעולה אחרונה: <b>{la_amt}</b> · {la_date}</div>'
        except: pass

    admin_badge = "<span style='background:rgba(255,255,255,0.15);border-radius:8px;padding:3px 10px;font-size:0.75rem;margin-right:8px'>מנהל מערכת</span>" if is_admin else ""
    gr1, gr2 = st.columns([5, 1])
    with gr1:
        st.markdown(f"""
        <div class="greeting-bar">
            <div>
                <div class="greeting-name">שלום, {u.get('שם משתמש','')} 👋 {admin_badge}</div>
                <div class="greeting-uid">מספר משתמש: {uid} · עודכן בשעה {sync_time}</div>
                {last_act_html}
            </div>
            <div class="greeting-badge">💰 מערכת פיננסית</div>
        </div>""", unsafe_allow_html=True)
    with gr2:
        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
        if st.button("🔄 רענן", key="dash_refresh", use_container_width=True, help="טען מחדש נתונים עדכניים מ-Sheets"):
            st.cache_data.clear()
            st.rerun()

    # FIX #4 – חישוב מחזור נכון
    try:    total_volume = pd.to_numeric(df_actions["סכום"], errors="coerce").sum()
    except: total_volume = 0

    # חישוב הכנסות/הוצאות אישיות
    if not my_act.empty and "סכום נטו" in my_act.columns:
        net_vals = pd.to_numeric(my_act["סכום נטו"], errors="coerce").fillna(0)
        personal_income = net_vals[net_vals > 0].sum()
        personal_expense = net_vals[net_vals < 0].abs().sum()
    else:
        personal_income = personal_expense = 0.0

    c1,c2,c3 = st.columns(3)
    with c1:
        st.markdown(f"""<div class="mc-wrap mc-blue">
            <div class="mc-icon">💰</div>
            <div class="mc-label">יתרה נוכחית</div>
            <div class="mc-value">₪{balance:,.2f}</div>
            <div class="mc-sub">⏱ מתעדכן כל 10 דקות</div>
        </div>""", unsafe_allow_html=True)
    with c2:
        if is_admin:
            st.markdown(f"""<div class="mc-wrap mc-green">
                <div class="mc-icon">🔄</div>
                <div class="mc-label">מחזור כללי במערכת</div>
                <div class="mc-value">₪{total_volume:,.0f}</div>
                <div class="mc-sub">סך כל הפעולות</div>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"""<div class="mc-wrap mc-green">
                <div class="mc-icon">📥</div>
                <div class="mc-label">סך הכנסות</div>
                <div class="mc-value">₪{personal_income:,.0f}</div>
                <div class="mc-sub">זכויות שהתקבלו</div>
            </div>""", unsafe_allow_html=True)
    with c3:
        if is_admin:
            st.markdown(f"""<div class="mc-wrap mc-purple">
                <div class="mc-icon">📋</div>
                <div class="mc-label">סך פעולות מערכת</div>
                <div class="mc-value">{len(df_actions)}</div>
                <div class="mc-sub">כלל הפעולות</div>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"""<div class="mc-wrap mc-purple">
                <div class="mc-icon">📤</div>
                <div class="mc-label">סך הוצאות</div>
                <div class="mc-value">₪{personal_expense:,.0f}</div>
                <div class="mc-sub">חובות ששולמו</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("")
    col_chart, col_feed = st.columns([1, 1.4])

    # FIX #3 – גרף עמודות עם ציר X תאריכים
    with col_chart:
        st.markdown('<div class="section-title">📊 גרף פעילות</div>', unsafe_allow_html=True)
        if not my_act.empty:
            view = st.radio("תקופה", ["החודש הנוכחי","כל הפעולות"], horizontal=True, key="dash_view")
            df_c = my_act.copy()
            if view == "החודש הנוכחי" and "תאריך" in df_c.columns:
                df_c = df_c[df_c["תאריך"].astype(str).str.startswith(datetime.now().strftime("%Y-%m"))]
            if not df_c.empty and "תאריך" in df_c.columns:
                try:
                    df_c["סכום נטו"] = pd.to_numeric(df_c["סכום נטו"], errors="coerce").fillna(0)
                    df_c["צבע"] = df_c["סכום נטו"].apply(lambda x: "הכנסה" if x>=0 else "הוצאה")
                    fig = px.bar(df_c.tail(20), x="תאריך", y="סכום נטו", color="צבע",
                                 color_discrete_map={"הכנסה":"#22c55e","הוצאה":"#ef4444"})
                    style_fig(fig, height=280)
                    fig.update_layout(xaxis_title="תאריך", yaxis_title="₪")
                    st.plotly_chart(fig, use_container_width=True)
                except Exception:
                    render_empty_state("📊", "אין מספיק נתונים לגרף")
            else:
                render_empty_state("📅", "אין פעולות בתקופה זו", "נסה לבחור ‘כל הפעולות’")
        else:
            render_empty_state("💳", "עדיין אין פעולות", "הפעולות שלך יופיעו כאן לאחר הביצוע הראשון")

    with col_feed:
        st.markdown('<div class="section-title">⚡ 5 פעולות אחרונות</div>', unsafe_allow_html=True)
        feed = df_actions if is_admin else my_act
        if not feed.empty:
            for _, row in feed.tail(5).iloc[::-1].iterrows():
                render_smart_card(row)
        else:
            render_empty_state("📥", "אין פעולות להצגה")


# ============================
# Tab 2 – עובר ושב  FIX #5
# ============================
def render_history(u, is_admin, df_users, df_actions):
    uid = str(u.get("מספר משתמש",""))

    if is_admin:
        my_act = df_actions.copy()
        try:
            my_act["כיוון"]    = "חובה"
            my_act["סכום נטו"] = pd.to_numeric(my_act["סכום"], errors="coerce").fillna(0)
            my_act["תיאור"]    = my_act.apply(lambda r: f"העברה מ-{r.get('שם מקור','').strip()} אל {r.get('שם יעד','').strip()}", axis=1)
        except Exception: pass
    else:
        my_act = process_user_actions(df_actions, uid)

    if my_act.empty:
        render_empty_state("📋", "אין פעולות להצגה", "כאשר יבוצעו פעולות הן יופיעו כאן")
        return

    for col in ["תאריך","שעה","סכום","סטטוס","כיוון","תיאור","שם מקור","שם יעד"]:
        if col not in my_act.columns: my_act[col] = ""

    my_act["סכום_num"] = pd.to_numeric(my_act["סכום"], errors="coerce").fillna(0)

    # גרפים
    st.markdown('<div class="section-title">📈 אנליטיקס</div>', unsafe_allow_html=True)
    tab_pie, tab_line = st.tabs(["🥧 לאן הכסף עבר?","📉 מגמת יתרה"])

    with tab_pie:
        try:
            debits = my_act[my_act["כיוון"]=="חובה"].copy()
            if not debits.empty and "שם יעד" in debits.columns:
                pie_df = debits.groupby("שם יעד")["סכום_num"].sum().reset_index()
                pie_df.columns = ["שם","סכום"]
                pie_df = pie_df[pie_df["שם"].str.strip()!=""]
                fig_pie = px.pie(pie_df, names="שם", values="סכום",
                                 title="פילוח הוצאות לפי יעד", hole=0.4,
                                 color_discrete_sequence=px.colors.sequential.Blues_r)
                style_fig(fig_pie, height=300)
                st.plotly_chart(fig_pie, use_container_width=True)
            else: render_empty_state("🦷", "אין נתוני הוצאות", "כאשר תבצע תשלומים, הגרף יופיע כאן")
        except Exception as e: st.warning(f"שגיאה: {e}")

    with tab_line:
        try:
            if "תאריך" in my_act.columns:
                line_df = my_act[["תאריך","סכום נטו"]].copy()
                line_df["סכום נטו"] = pd.to_numeric(line_df["סכום נטו"], errors="coerce").fillna(0)
                line_df = line_df.sort_values("תאריך")
                line_df["יתרה מצטברת"] = line_df["סכום נטו"].cumsum()
                fig_l = px.line(line_df, x="תאריך", y="יתרה מצטברת",
                                markers=True, color_discrete_sequence=["#0f3460"])
                style_fig(fig_l, height=300)
                fig_l.update_layout(yaxis_title="₪")
                st.plotly_chart(fig_l, use_container_width=True)
            else: render_empty_state("📉", "אין מספיק נתונים", "נדרשות לפחות 2 פעולות עם תאריכים")
        except Exception: render_empty_state("📉", "לא היה ניתן לטעון את הגרף")

    st.divider()

    # סינון
    st.markdown('<div class="section-title">🔍 סינון וחיפוש</div>', unsafe_allow_html=True)
    st.markdown("""
    <style>.filter-card{background:white;border-radius:16px;padding:20px 24px 8px;
    box-shadow:0 2px 12px rgba(0,0,0,0.06);margin-bottom:16px}</style>
    <div class="filter-card">""", unsafe_allow_html=True)
    fc1,fc2,fc3 = st.columns([2,2,2])
    with fc1: search_text = st.text_input("🔎 חיפוש חופשי","", key="hist_search")
    with fc2:
        date_filter = st.selectbox("📅 תקופה",["הכל","החודש הנוכחי","חודש קודם","טווח מותאם"], key="hist_date")
        date_from = date_to = None
        if date_filter=="טווח מותאם":
            d1c,d2c = st.columns(2)
            with d1c: date_from = st.date_input("מ-", key="hist_df")
            with d2c: date_to   = st.date_input("עד", key="hist_dt")
    with fc3:
        status_filter = st.multiselect("סטטוס",["הכל","זכות","חובה","כושלות","פעולות מנהל"],
                                       default=["הכל"], key="hist_status")
    fc4,fc5 = st.columns(2)
    with fc4: min_amt = st.number_input("סכום מינימלי ₪", value=0.0, step=10.0, key="hist_min")
    with fc5: max_amt = st.number_input("סכום מקסימלי ₪ (ביטול = אין גבול)", value=0.0, step=100.0, key="hist_max", help="0 = ללא גבול")

    # החלת סינונים
    filtered = my_act.copy()
    if search_text:
        filtered = filtered[filtered.apply(lambda r: search_text.lower() in " ".join(r.astype(str).values).lower(), axis=1)]

    if date_filter == "החודש הנוכחי":
        filtered = filtered[filtered["תאריך"].astype(str).str.startswith(datetime.now().strftime("%Y-%m"))]
    elif date_filter == "חודש קודם":
        lm = (datetime.now().replace(day=1) - pd.Timedelta(days=1)).strftime("%Y-%m")
        filtered = filtered[filtered["תאריך"].astype(str).str.startswith(lm)]
    elif date_filter == "טווח מותאם" and date_from and date_to:
        filtered = filtered[(filtered["תאריך"].astype(str)>=str(date_from)) & (filtered["תאריך"].astype(str)<=str(date_to))]

    if status_filter and "הכל" not in status_filter:
        conds = []
        if "זכות"        in status_filter: conds.append(filtered["כיוון"]=="זכות")
        if "חובה"        in status_filter: conds.append(filtered["כיוון"]=="חובה")
        if "כושלות"      in status_filter: conds.append(filtered["סטטוס"].astype(str).str.contains("כושל|נכשל",na=False))
        if "פעולות מנהל" in status_filter: conds.append(filtered["סטטוס"].astype(str).str.contains("מנהל",na=False))
        if conds: filtered = filtered[functools.reduce(lambda a,b: a|b, conds)]

    filtered = filtered[(filtered["סכום_num"].abs()>=min_amt) & (max_amt == 0.0 or filtered["סכום_num"].abs()<=max_amt)]
    st.markdown('</div>', unsafe_allow_html=True)  # close filter-card

    # מיון
    sort_options = ["תאריך ↓ (חדש ראשון)","תאריך ↑ (ישן ראשון)","סכום ↓ (גדול ראשון)","סכום ↑ (קטן ראשון)","זכויות קודם","חובות קודם"]
    res1, res2 = st.columns([3, 1])
    with res1:
        st.markdown(f"""
        <div style="background:white;border-radius:12px;padding:10px 18px;
                    box-shadow:0 1px 6px rgba(0,0,0,0.05);margin-bottom:4px;height:42px;
                    display:flex;align-items:center">
            <span style="font-size:0.85rem;color:#475569;font-weight:600">
                📋 נמצאו <span style="color:#0f3460;font-size:1rem">{len(filtered)}</span> פעולות
            </span>
        </div>""", unsafe_allow_html=True)
    with res2:
        sort_opt = st.selectbox("מיון", sort_options, key="hist_sort", label_visibility="collapsed")

    # החלת מיון
    if "תאריך ↑" in sort_opt:
        filtered = filtered.sort_values("תאריך", ascending=True)
    elif "תאריך ↓" in sort_opt:
        filtered = filtered.sort_values("תאריך", ascending=False)
    elif "סכום ↓" in sort_opt:
        filtered = filtered.assign(_abs=filtered["סכום_num"].abs()).sort_values("_abs", ascending=False).drop(columns="_abs")
    elif "סכום ↑" in sort_opt:
        filtered = filtered.assign(_abs=filtered["סכום_num"].abs()).sort_values("_abs", ascending=True).drop(columns="_abs")
    elif "זכויות" in sort_opt:
        filtered = filtered.sort_values("כיוון", ascending=False) if "כיוון" in filtered.columns else filtered
    elif "חובות" in sort_opt:
        filtered = filtered.sort_values("כיוון", ascending=True) if "כיוון" in filtered.columns else filtered

    # FIX #5 – reset page on filter/sort change
    fkey = f"{search_text}|{date_filter}|{status_filter}|{min_amt}|{max_amt}|{sort_opt}"
    if st.session_state.get("_hist_fkey") != fkey:
        st.session_state._hist_fkey = fkey
        st.session_state.hist_page  = 0

    # ייצוא
    ex1,ex2,_ = st.columns([1,1,4])
    with ex1:
        try:
            buf = BytesIO()
            exp_df = filtered.drop(columns=["סכום_num"],errors="ignore")
            user_name = u.get("שם משתמש", "") if 'u' in dir() else ""
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                # Write data starting from row 4 (leave space for header)
                exp_df.to_excel(writer, index=False, startrow=3, sheet_name="פעולות")
                ws = writer.sheets["פעולות"]
                # Header rows
                ws["A1"] = "💰 שיעבודא פון – דו"ח פעולות"
                ws["A2"] = f"משתמש: {user_name}  |  תאריך הפקה: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  סך פעולות: {len(exp_df)}"
                # Style header
                from openpyxl.styles import Font, PatternFill, Alignment
                ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F3460")
                ws["A2"].font = Font(name="Arial", size=10, color="475569")
                ws["A1"].alignment = Alignment(horizontal="right")
                ws["A2"].alignment = Alignment(horizontal="right")
                # Column headers style (row 4)
                hdr_fill = PatternFill("solid", fgColor="0F3460")
                hdr_font = Font(name="Arial", bold=True, color="FFFFFF")
                for cell in ws[4]:
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal="right")
                # Auto-width columns
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=8)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
            st.download_button("📥 ייצוא Excel", data=buf.getvalue(), file_name="shiabudefon.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception: st.warning("התקן openpyxl")
    with ex2:
        try:
            exp = filtered.drop(columns=["סכום_num"],errors="ignore")
            trs = "".join(
                f"<tr style='background:{'#fff5f5' if r.get('כיוון')=='חובה' else '#f0fff4'}'>" +
                "".join(f"<td>{v}</td>" for v in r.values) + "</tr>"
                for _,r in exp.iterrows()
            )
            ths = "".join(f"<th>{c}</th>" for c in exp.columns)
            html = (f'<!DOCTYPE html><html dir="rtl"><head><meta charset="UTF-8">'
                    f'<style>body{{font-family:Arial;direction:rtl;padding:24px}}'
                    f'h2{{color:#0f3460}}table{{border-collapse:collapse;width:100%;font-size:12px}}'
                    f'th{{background:#0f3460;color:white;padding:8px;border:1px solid #ccc}}'
                    f'td{{padding:6px;border:1px solid #ddd}}'
                    f'@media print{{@page{{size:A4 landscape;margin:1cm}}button{{display:none}}}}'
                    f'</style></head><body>'
                    f'<h2>שיעבודא פון – דו"ח פעולות</h2>'
                    f'<p>תאריך הפקה: {datetime.now().strftime("%d/%m/%Y %H:%M")} | פעולות: {len(filtered)}</p>'
                    f'<table><thead><tr>{ths}</tr></thead><tbody>{trs}</tbody></table>'
                    f'<br><button onclick="window.print()">🖨️ הדפס / שמור כ-PDF</button>'
                    f'</body></html>')
            st.download_button("📄 PDF (הדפסה)", data=html.encode("utf-8"),
                               file_name="shiabudefon_report.html", mime="text/html")
        except Exception: pass

    # דפדוף
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    PAGE_SIZE   = 20
    total_pages = max(1,(len(filtered)+PAGE_SIZE-1)//PAGE_SIZE)
    if "hist_page" not in st.session_state: st.session_state.hist_page = 0
    st.session_state.hist_page = min(st.session_state.hist_page, total_pages-1)

    pg1,pg2,pg3 = st.columns([1,3,1])
    with pg1:
        if st.session_state.hist_page>0:
            if st.button("◀ הקודם", key="pg_prev", use_container_width=True):
                st.session_state.hist_page-=1; st.rerun()
    with pg2:
        progress = (st.session_state.hist_page+1) / total_pages
        st.markdown(f"""
        <div class="pg-info">
            עמוד <span style="color:#0f3460;font-size:1.05rem;font-weight:800">
            {st.session_state.hist_page+1}</span>
            <span style="color:#cbd5e1"> / {total_pages}</span>
            <div style="background:#e2e8f0;border-radius:4px;height:4px;margin-top:8px">
                <div style="background:linear-gradient(90deg,#0f3460,#1565c0);height:4px;
                            border-radius:4px;width:{progress*100:.0f}%"></div>
            </div>
        </div>""", unsafe_allow_html=True)
    with pg3:
        if st.session_state.hist_page<total_pages-1:
            if st.button("הבא ▶", key="pg_next", use_container_width=True):
                st.session_state.hist_page+=1; st.rerun()

    start   = st.session_state.hist_page * PAGE_SIZE
    page_df = filtered.iloc[::-1].reset_index(drop=True).iloc[start:start+PAGE_SIZE]

    if filtered.empty:
        render_empty_state("🔍", "לא נמצאו פעולות התואמות את החיפוש שלך", "נסה לשנות את קריתריוני הסינון"); return
    if not filtered.empty and "סכום_num" in filtered.columns:
        f_credits  = filtered[filtered.get("כיוון", pd.Series()) == "זכות"]["סכום_num"].abs().sum() if "כיוון" in filtered.columns else 0
        f_debits   = filtered[filtered.get("כיוון", pd.Series()) == "חובה"]["סכום_num"].abs().sum() if "כיוון" in filtered.columns else 0
        f_credits_v = filtered[filtered["סכום_num"] > 0]["סכום_num"].sum()
        f_debits_v  = filtered[filtered["סכום_num"] < 0]["סכום_num"].abs().sum()
        total_in  = f_credits_v or f_credits
        total_out = f_debits_v  or f_debits
        if total_in or total_out:
            st.markdown(f"""
            <div style="display:flex;gap:12px;margin-bottom:14px;flex-wrap:wrap">
                <div style="flex:1;min-width:140px;background:#f0fdf4;border-radius:12px;
                            padding:12px 16px;border:1px solid #dcfce7;text-align:center">
                    <div style="font-size:0.75rem;color:#16a34a;font-weight:600;margin-bottom:3px">📥 סך הכנסות בסינון</div>
                    <div style="font-size:1.15rem;font-weight:800;color:#14532d;direction:ltr">₪{total_in:,.2f}</div>
                </div>
                <div style="flex:1;min-width:140px;background:#fff5f5;border-radius:12px;
                            padding:12px 16px;border:1px solid #fee2e2;text-align:center">
                    <div style="font-size:0.75rem;color:#dc2626;font-weight:600;margin-bottom:3px">📤 סך הוצאות בסינון</div>
                    <div style="font-size:1.15rem;font-weight:800;color:#7f1d1d;direction:ltr">₪{total_out:,.2f}</div>
                </div>
                <div style="flex:1;min-width:140px;background:#eff6ff;border-radius:12px;
                            padding:12px 16px;border:1px solid #dbeafe;text-align:center">
                    <div style="font-size:0.75rem;color:#2563eb;font-weight:600;margin-bottom:3px">📊 נטו</div>
                    <div style="font-size:1.15rem;font-weight:800;color:#1e3a8a;direction:ltr">₪{total_in-total_out:,.2f}</div>
                </div>
            </div>""", unsafe_allow_html=True)

    for _,row in page_df.iterrows():
        render_smart_card(row, highlight=search_text)


# ============================
# Tab 3 – פרטים אישיים
# ============================
def render_personal(u, df_users):
    uid = str(u.get("מספר משתמש",""))
    df  = df_users.copy()
    df["מספר משתמש"] = df["מספר משתמש"].astype(str).str.strip()
    row       = df[df["מספר משתמש"]==uid]
    user_data = row.iloc[0].to_dict() if not row.empty else u

    st.markdown('<div class="section-title">👤 החשבון שלי – פרטים מזהים</div>', unsafe_allow_html=True)
    raw_pwd = str(user_data.get("סיסמה", "—"))
    masked_pwd = ("*" * len(raw_pwd)) if raw_pwd not in ("—", "", "nan") else "—"
    show_p = st.session_state.get("_show_personal_pwd", False)
    pwd_display = raw_pwd if show_p else masked_pwd
    fields = [("שם משתמש","שם משתמש",None),("מספר משתמש","מספר משתמש",None),
              ("תעודת זהות","תעודת זהות",None),("כתובת","כתובת",None),
              ("סיסמה נוכחית","סיסמה",pwd_display)]
    rows_html = "".join(
        f'<div class="detail-row"><span class="detail-label">{lbl}:</span>'
        f'<span class="detail-value">{override if override is not None else user_data.get(key,"—")}</span></div>'
        for lbl,key,override in fields
    )
    st.markdown(f"""<div class="detail-block">{rows_html}
        <div class="readonly-note">⚠️ המידע מוצג לקריאה בלבד. לעדכון פרטים, חייג למערכת הטלפונית.</div>
    </div>""", unsafe_allow_html=True)
    eye_lbl = "🙈 הסתר סיסמה" if show_p else "👁 הצג סיסמה"
    if st.button(eye_lbl, key="toggle_personal_pwd"):
        st.session_state._show_personal_pwd = not show_p
        st.rerun()

    st.markdown('<div class="section-title">📱 רשימת צינתוקים</div>', unsafe_allow_html=True)
    raw    = user_data.get("צינתוקים", user_data.get("טלפונים",""))
    phones = [p.strip() for p in str(raw).split(",") if p.strip() and str(raw) not in ("","nan","None")]
    tags   = "".join(f'<span class="phone-tag">📞 {p}</span>' for p in phones) if phones else '<span style="color:#aaa">לא רשומים מספרים</span>'
    cnt    = f"רשומים <b>{len(phones)}</b> מספרים." if phones else "לא רשומים מספרים."
    st.markdown(f"""<div class="detail-block">
        <div style="margin-bottom:12px;color:#555">{cnt}</div>
        <div>{tags}</div>
        <div class="readonly-note">➕ להוספה/הסרה, חייג למערכת הטלפונית.</div>
    </div>""", unsafe_allow_html=True)


# ============================
# Tab 4 – צ'אט AI
# ============================
def render_chat_tab(u, is_admin, df_users, df_actions):
    uid = str(u.get("מספר משתמש",""))
    if "messages" not in st.session_state: st.session_state.messages = []

    if not st.session_state.messages:
        st.markdown("""
        <div style="background:linear-gradient(135deg,#eff6ff,#dbeafe);border-radius:18px;
                    padding:26px 28px;margin-bottom:20px;border:1px solid #bfdbfe;direction:rtl">
            <div style="font-size:1.6rem;margin-bottom:8px">🤖</div>
            <div style="font-weight:700;color:#1e3a8a;font-size:1rem;margin-bottom:6px">בוט שיעבודא פון AI</div>
            <div style="color:#3b82f6;font-size:0.86rem;line-height:1.7">
                שאל אותי כל שאלה על הפעולות שלך:
                <span style="background:#dbeafe;border-radius:8px;padding:2px 9px;margin:3px;display:inline-block">סכם הוצאות החודש</span>
                <span style="background:#dbeafe;border-radius:8px;padding:2px 9px;margin:3px;display:inline-block">מי קיבל הכי הרבה?</span>
                <span style="background:#dbeafe;border-radius:8px;padding:2px 9px;margin:3px;display:inline-block">תסכם לי את הפעולות</span>
            </div>
        </div>""", unsafe_allow_html=True)

    for msg in st.session_state.messages:
        with st.chat_message(msg["role"]):
            st.write(msg["content"])
            # token count shown to admins only
            if msg.get("tokens") and is_admin:
                st.caption(f"🪙 {msg['tokens']}")

    if prompt := st.chat_input("שאל אותי (לדוגמה: תסכם לי הוצאות)..."):
        st.session_state.messages.append({"role":"user","content":prompt})
        with st.chat_message("user"): st.write(prompt)
        with st.chat_message("assistant"):
            with st.spinner("חושב..."):
                if is_admin:
                    context = f"פעולות:\n{df_actions.to_csv()}\nמשתמשים (ללא סיסמאות):\n{df_users.drop(columns=['סיסמה'], errors='ignore').to_csv()}"
                else:
                    my_act   = process_user_actions(df_actions, uid)
                    curr_row = df_users[df_users["מספר משתמש"].astype(str)==uid].drop(columns=["סיסמה"], errors="ignore")
                    context  = f"פרטים:\n{curr_row.to_csv()}\nפעולות:\n{my_act.drop(columns=['סכום_num'], errors='ignore').to_csv()}"
                reply = "מצטער, לא הצלחתי לקבל תשובה."; tokens_info = ""
                model_name = _get_gemini_model_name()
                for key in get_api_keys():
                    try:
                        genai.configure(api_key=key)
                        res = genai.GenerativeModel(model_name).generate_content(
                            f"{SYSTEM_MANUAL}\n{context}\nשאלה: {prompt}",
                            generation_config=genai.types.GenerationConfig(max_output_tokens=400))
                        reply = res.text
                        try: tokens_info = f"טוקנים: {res.usage_metadata.total_token_count}"
                        except Exception: pass
                        break
                    except Exception: continue
                st.write(reply)
                if tokens_info and is_admin: st.caption(f"🪙 {tokens_info}")
        st.session_state.messages.append({"role":"assistant","content":reply,"tokens":tokens_info})


# ============================
# Admin
# ============================
def render_admin(df_users, df_actions):
    try:    total_vol = pd.to_numeric(df_actions["סכום"], errors="coerce").sum()
    except: total_vol = 0
    try:    avg_tx = total_vol / max(len(df_actions), 1)
    except: avg_tx = 0

    st.markdown('<div class="section-title">🛠️ לוח בקרה – מנהל מערכת</div>', unsafe_allow_html=True)
    c1,c2,c3,c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="mc-wrap mc-blue"><div class="mc-icon">👥</div><div class="mc-label">משתמשים רשומים</div><div class="mc-value">{len(df_users)}</div><div class="mc-sub">סך משתמשים</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="mc-wrap mc-green"><div class="mc-icon">📋</div><div class="mc-label">סך פעולות</div><div class="mc-value">{len(df_actions)}</div><div class="mc-sub">פעולות שבוצעו</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="mc-wrap mc-purple"><div class="mc-icon">💸</div><div class="mc-label">מחזור כללי</div><div class="mc-value">₪{total_vol:,.0f}</div><div class="mc-sub">סך כל העברות</div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="mc-wrap" style="background:linear-gradient(135deg,#b45309,#d97706);box-shadow:0 6px 20px rgba(180,83,9,0.35)"><div class="mc-icon">📊</div><div class="mc-label">עסקה ממוצעת</div><div class="mc-value">₪{avg_tx:,.0f}</div><div class="mc-sub">לפעולה</div></div>', unsafe_allow_html=True)

    st.markdown("")
    st.markdown("""
    <style>
    [data-testid="stDataFrame"] { border-radius: 14px !important; overflow: hidden !important;
        box-shadow: 0 2px 12px rgba(0,0,0,0.06) !important; }
    </style>""", unsafe_allow_html=True)

    at1,at2 = st.tabs(["📋 כל הפעולות","👥 משתמשים"])
    with at1:
        st.markdown(f'<div style="color:#64748b;font-size:0.8rem;margin-bottom:8px">50 פעולות אחרונות מתוך {len(df_actions)}</div>', unsafe_allow_html=True)
        st.dataframe(df_actions.tail(50).iloc[::-1], hide_index=True, use_container_width=True)
    with at2:
        st.markdown(f'<div style="color:#64748b;font-size:0.8rem;margin-bottom:8px">{len(df_users)} משתמשים רשומים</div>', unsafe_allow_html=True)
        st.dataframe(df_users, hide_index=True, use_container_width=True)


# ============================
# דפי CMS
# ============================
def render_info_page(title: str, content: str):
    st.markdown(f"""<div class="info-page">
        <h2>{title}</h2>
        <div style="line-height:1.9">{content.replace(chr(10),'<br>')}</div>
    </div>""", unsafe_allow_html=True)


# ============================
# MAIN
# ============================
def main():
    SESSION_TIMEOUT = 15 * 60  # 15 דקות

    if "page"          not in st.session_state: st.session_state.page = "login"
    if "authenticated" not in st.session_state: st.session_state.authenticated = False

    # Handle ?page= query params from hamburger menu links
    try:
        qp = st.query_params.get("page", None)
        if qp and qp in ("app", "about", "contact", "terms", "login"):
            st.session_state.page = qp
            st.query_params.clear()
    except Exception:
        pass

    # Session timeout check
    if st.session_state.authenticated:
        now = datetime.now().timestamp()
        last = st.session_state.get("_last_activity", now)
        if now - last > SESSION_TIMEOUT:
            st.session_state.authenticated = False
            st.session_state.page = "login"
            st.session_state._timeout_msg = True
        else:
            st.session_state._last_activity = now

    page = st.session_state.page
    if st.session_state.authenticated and page=="login":
        st.session_state.page = "app"; page = "app"

    render_header(page)

    if page in ("about","contact","terms"):
        cms = get_cms_content()
        render_info_page({"about":"קצת עלינו","contact":"צור קשר","terms":"תקנון"}[page],
                         cms.get({"about":"קצת עלינו","contact":"צור קשר","terms":"תקנון"}[page],""))
        return

    if page=="login" or not st.session_state.authenticated:
        if st.session_state.pop("_timeout_msg", False):
            st.warning("⏰ נותקת עקב חוסר פעילות. אנא התחבר מחדש.")
        render_login_page(); return

    u        = st.session_state.user
    is_admin = st.session_state.get("is_admin", False)

    # FIX #13 – spinner טעינה ראשונית
    with st.spinner("⏳ טוען נתונים..."):
        df_users, df_actions, _ = get_all_data()
    st.session_state._last_sync = datetime.now().strftime("%H:%M:%S")

    # FIX #6 – כפתור צף אמיתי
    inject_chat_fab_js()

    tabs_labels = ["🏠 דשבורד","📊 עובר ושב","👤 פרטים אישיים","🤖 צ'אט AI"]
    if is_admin: tabs_labels.append("🛠️ ניהול מנהל")
    tab_objs = st.tabs(tabs_labels)

    with tab_objs[0]: render_dashboard(u, is_admin, df_users, df_actions)
    with tab_objs[1]: render_history(u, is_admin, df_users, df_actions)
    with tab_objs[2]: render_personal(u, df_users)
    with tab_objs[3]: render_chat_tab(u, is_admin, df_users, df_actions)
    if is_admin:
        with tab_objs[4]: render_admin(df_users, df_actions)

    # Footer
    st.markdown("""
    <div style="text-align:center;padding:28px 0 12px;color:#cbd5e1;font-size:0.78rem;border-top:1px solid #e2e8f0;margin-top:24px">
        💰 <b style="color:#94a3b8">שיעבודא פון</b> &nbsp;&middot;&nbsp; מערכת פיננסית מתקדמת &nbsp;&middot;&nbsp;
        כל הזכויות שמורות &copy; {datetime.now().year}
    </div>""", unsafe_allow_html=True)


main()
